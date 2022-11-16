from pathlib import Path
from typing import Dict, Any, List

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from django_scaffolding_tools._experimental.general_ledger.parsers import get_default_mappings


def write_headers(sheet: Worksheet, headers: List[str]):
    for i, header in enumerate(headers, 1):
        sheet.merge_cells(f'A{i}:L{i}')
        cell = sheet.cell(row=i, column=1)
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center')


def write_transactions(target_file: Path, parsed_results: Dict[str, Any], start_row: int = 6,
                       column_mappings: Dict[int, Dict[str, Any]] = get_default_mappings()):
    if target_file.exists():
        wb = load_workbook(target_file)
    else:
        wb = Workbook()
    for account_id in parsed_results['accounts'].keys():
        sheet = wb.create_sheet(account_id)
        # Write sheet headers
        write_headers(sheet, parsed_results['headers'])
        # Setting column widths
        for col_num, column_mapping in column_mappings.items():
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = column_mapping.get('width', 20)
        # Write column headers
        for col_num, col_mapping in column_mappings.items():
            cell = sheet.cell(column=col_num, row=start_row - 1)
            cell.value = col_mapping['title']
            cell.font = Font(bold=True)
        # Write transactions
        row = start_row
        transaction_list = parsed_results['accounts'][account_id]
        for transaction in transaction_list:
            for col_num, col_mapping in column_mappings.items():
                sheet.cell(column=col_num, row=row, value=transaction[col_mapping['name']])
                if col_mapping.get('number_format') is not None:
                    sheet.cell(column=col_num, row=row).number_format = col_mapping.get('number_format')
            row += 1
    wb.save(target_file)
