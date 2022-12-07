from pathlib import Path
from typing import List, Any, Dict, Protocol

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

COLUMN_MAPPINGS = {
    1: {'name': 'account_id', 'title': 'Account ID', 'width': 12},
    2: {'name': 'account_description', 'title': 'Account Description', 'width': 24},
    3: {'name': 'date', 'title': 'Date', 'number_format': 'DD/MM/YYYY', 'width': 12},
    4: {'name': 'reference', 'title': 'Reference', 'width': 30},
    5: {'name': 'journal', 'title': 'Jrnl', 'width': 12},
    6: {'name': 'description', 'title': 'Trans Description', 'width': 36},
    7: {'name': 'debit_amount', 'title': 'Debit Amt', 'number_format': '#,##0.00', 'width': 12},
    8: {'name': 'credit_amount', 'title': 'Credit Amt', 'number_format': '#,##0.00', 'width': 12},
    9: {'name': 'balance', 'title': 'Balance', 'number_format': '#,##0.00', 'width': 12},
}

COLUMN_MAPPINGS_2 = {
    '1': {'name': 'account_id', 'title': 'Account ID', 'width': 12},
    '2': {'name': 'account_description', 'title': 'Account Description', 'width': 24},
    '3': {'name': 'date', 'title': 'Date', 'number_format': 'DD/MM/YYYY', 'width': 12},
    '4': {'name': 'reference', 'title': 'Reference', 'width': 30},
    '5': {'name': 'journal', 'title': 'Jrnl', 'width': 12},
    '6': {'name': 'description', 'title': 'Trans Description', 'width': 36},
    '7': {'name': 'debit_amount', 'title': 'Debit Amt', 'number_format': '#,##0.00', 'width': 12},
    '8': {'name': 'credit_amount', 'title': 'Credit Amt', 'number_format': '#,##0.00', 'width': 12},
    '9': {'name': 'balance', 'title': 'Balance', 'number_format': '#,##0.00', 'width': 12},
}


def get_default_mappings():
    return COLUMN_MAPPINGS


class ParsingConfiguration(Protocol):
    @property
    def start_row(self) -> int:
        ...

    @property
    def sheet_name(self) -> str:
        ...

    @property
    def column_mappings(self) -> Dict[str, Dict[str, Any]]:
        ...


def parse_general_ledger(general_ledger_file: Path, start_row: int = 6,
                         sheet_name: str = 'General Ledger',
                         column_mappings: Dict[int, Dict[str, Any]] = None,
                         ) -> Dict[str, Any]:
    if column_mappings is None:
        column_mappings = get_default_mappings()
    accounts = dict()
    headers = list()
    wb = load_workbook(general_ledger_file)
    sheet = wb[sheet_name]
    # Read headers
    for row in range(1, 5):
        headers.append(sheet.cell(row=row, column=1).value)
    # Read account transactions
    last_row = sheet.max_row + 1
    current_account = None
    for row in tqdm(range(start_row, last_row)):
        transaction_dict = dict()
        account_id = sheet.cell(row=row, column=1).value
        if current_account is None:
            current_account = account_id
            accounts[account_id] = list()
        elif account_id == '':
            pass
        elif current_account != account_id:
            current_account = account_id
            accounts[account_id] = list()

        for col, col_value in column_mappings.items():
            cell_obj = sheet.cell(row=row, column=col)
            value = cell_obj.value
            transaction_dict[col_value['name']] = value
        accounts[current_account].append(transaction_dict)
    parsed_results = {'headers': headers, 'accounts': accounts}
    return parsed_results
